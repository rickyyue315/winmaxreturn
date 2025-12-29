import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

# 設定頁面配置
st.set_page_config(
    page_title="退貨建議分析系統",
    page_icon="📦",
    layout="wide"
)

def convert_to_string_format(value):
    """確保 Article 欄位為 12 位字符串格式"""
    if pd.isna(value):
        return ""
    value_str = str(value).strip()
    # 移除小數點（如果是浮點數）
    if '.' in value_str:
        value_str = value_str.split('.')[0]
    # 確保是 12 位數字
    if value_str.isdigit() and len(value_str) <= 12:
        return value_str.zfill(12)
    return value_str

def safe_convert_to_int(value):
    """安全轉換為整數，異常值設為 0"""
    try:
        if pd.isna(value):
            return 0
        return max(0, int(float(value)))
    except:
        return 0

def safe_convert_to_string(value):
    """安全轉換為字符串"""
    if pd.isna(value):
        return ""
    return str(value).strip()

def preprocess_data(df):
    """數據預處理與驗證"""
    df_processed = df.copy()
    
    # 確保 Article 欄位為 12 位字符串格式
    df_processed['Article'] = df_processed['Article'].apply(convert_to_string_format)
    
    # 字符串欄位處理
    string_columns = ['OM', 'RP Type', 'Site']
    for col in string_columns:
        if col in df_processed.columns:
            df_processed[col] = df_processed[col].apply(safe_convert_to_string)
    
    # 整數欄位處理
    int_columns = ['SaSa Net Stock', 'Pending Received', 'Safety Stock', 'Last Month Sold Qty', 'MTD Sold Qty']
    for col in int_columns:
        if col in df_processed.columns:
            df_processed[col] = df_processed[col].apply(safe_convert_to_int)
    
    # 銷量異常值校正
    df_processed['Notes'] = ""
    
    for idx, row in df_processed.iterrows():
        notes = []
        
        # 檢查銷量數據範圍
        for col in ['Last Month Sold Qty', 'MTD Sold Qty']:
            if col in df_processed.columns:
                if row[col] > 100000:
                    df_processed.loc[idx, col] = 100000
                    notes.append(f'{col}銷量數據超出範圍')
        
        df_processed.loc[idx, 'Notes'] = '; '.join(notes)
    
    return df_processed

def calculate_effective_sold_qty(row):
    """計算有效銷量"""
    last_month = row.get('Last Month Sold Qty', 0)
    mtd = row.get('MTD Sold Qty', 0)
    
    # 優先使用 Last Month Sold Qty，若為 0 則使用 MTD Sold Qty
    if last_month > 0:
        return last_month
    else:
        return mtd

def get_top20_percent_threshold(df, article):
    """計算該 Article 的銷量前 20% 門檻"""
    article_data = df[df['Article'] == article]
    if article_data.empty:
        return float('inf')
    
    sold_quantities = []
    for _, row in article_data.iterrows():
        effective_qty = calculate_effective_sold_qty(row)
        sold_quantities.append(effective_qty)
    
    if not sold_quantities:
        return float('inf')
    
    # 計算 80% 分位數（前 20% 的門檻）
    threshold = np.percentile(sold_quantities, 80)
    return threshold

def generate_return_recommendations(df, calculation_type="both"):
    """生成退貨建議
    
    Args:
        df: 數據框架
        calculation_type: 計算類型 ('nd_only', 'rf_only', 'both')
    """
    recommendations = []
    
    for _, row in df.iterrows():
        article = row['Article']
        om = row['OM']
        rp_type = row['RP Type']
        site = row['Site']
        net_stock = row['SaSa Net Stock']
        pending_received = row['Pending Received']
        safety_stock = row['Safety Stock']
        last_month_sold = row.get('Last Month Sold Qty', 0)
        mtd_sold = row.get('MTD Sold Qty', 0)
        product_hierarchy = row.get('Product Hierarchy', '')
        
        effective_sold_qty = calculate_effective_sold_qty(row)
        
        # ND 類型退倉
        if rp_type == "ND" and net_stock > 0 and calculation_type in ["nd_only", "both"]:
            return_qty = net_stock
            remaining_stock = net_stock - return_qty
            
            notes_parts = ['ND類型退倉']
            
            # 檢查是否有銷售記錄
            if effective_sold_qty > 0:
                notes_parts.append('曾有銷售記錄, Buyer需要留意是否需轉成RF及設定Safety')
            
            if row.get("Notes"):
                notes_parts.append(row.get("Notes"))
            
            recommendations.append({
                'Article': article,
                'Product Desc': row.get('Article Description', ''),
                'Product Hierarchy': product_hierarchy,
                'OM': om,
                'Return Site': site,
                'Receive Site': 'D001',
                'Return Qty': return_qty,
                'RP Type': rp_type,
                'Stock Qty': net_stock,
                'Safety Qty': safety_stock,
                'Last Month Sold Qty': last_month_sold,
                'MTD Sold Qty': mtd_sold,
                'Remaining Stock After Return': remaining_stock,
                'Notes': '; '.join(notes_parts),
                'Type': 'ND'
            })
        
        # RF 類型過剩退倉
        elif rp_type == "RF" and calculation_type in ["rf_only", "both"]:
            total_available = net_stock + pending_received
            
            # 檢查條件：庫存充足
            if total_available > safety_stock:
                # 檢查銷量是否非最高 20%
                top20_threshold = get_top20_percent_threshold(df, article)
                
                if effective_sold_qty < top20_threshold:
                    # 根據銷售量調整退貨後淨餘數量要求
                    # 若上月銷售量/MTD銷售量 其中一個月 > Safety Qty：退貨後淨餘數量需高於 Safety Qty 的 25% / +2件
                    # 若上月銷售量/MTD銷售量 同樣地 ≤ Safety Qty：退貨後淨餘數量只需高於 Safety Qty 1 件
                    if last_month_sold > safety_stock or mtd_sold > safety_stock:
                        # 銷售量高於 Safety Qty：退貨後淨餘數量需高於 Safety Qty 的 25% 且至少 +2 件
                        min_remaining_25_percent = int(safety_stock * 1.25)
                        min_remaining_2_more = safety_stock + 2
                        min_remaining = max(min_remaining_25_percent, min_remaining_2_more, 0)
                    else:
                        # 銷售量少於 Safety Qty：退貨後淨餘數量只需要高於 Safety Qty 1 件
                        min_remaining = max(safety_stock + 1, 0)
                    
                    # 計算可退貨數量
                    potential_return = total_available - safety_stock
                    max_return = total_available - min_remaining
                    
                    # 最終退貨數量（至少 2 件）
                    return_qty = min(potential_return, max_return)
                    
                    if return_qty >= 2 and return_qty <= net_stock:
                        remaining_stock = net_stock - return_qty
                        
                        notes_parts = ['RF類型過剩退倉']
                        if row.get("Notes"):
                            notes_parts.append(row.get("Notes"))
                        
                        recommendations.append({
                            'Article': article,
                            'Product Desc': row.get('Article Description', ''),
                            'Product Hierarchy': product_hierarchy,
                            'OM': om,
                            'Return Site': site,
                            'Receive Site': 'D001',
                            'Return Qty': return_qty,
                            'RP Type': rp_type,
                            'Stock Qty': net_stock,
                            'Safety Qty': safety_stock,
                            'Last Month Sold Qty': last_month_sold,
                            'MTD Sold Qty': mtd_sold,
                            'Remaining Stock After Return': remaining_stock,
                            'Notes': '; '.join(notes_parts),
                            'Type': 'RF'
                        })
    
    return pd.DataFrame(recommendations)

def create_excel_report(recommendations_df, df_original, calculation_type="both"):
    """創建 Excel 報告
    
    Args:
        recommendations_df: 退貨建議數據框架
        df_original: 原始數據框架
        calculation_type: 計算類型 ('nd_only', 'rf_only', 'both')
    """
    # 創建工作簿
    wb = Workbook()
    
    # 定義樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 工作表 1: 退貨建議
    ws1 = wb.active
    ws1.title = "退貨建議"
    
    # 寫入標題行
    headers = ['Product Hierarchy', 'Article', 'Product Desc', 'OM', 'Return Site', 'Receive Site', 'Return Qty',
               'RP Type', 'Stock Qty', 'Safety Qty', 'Last Month Sold Qty', 'MTD Sold Qty',
               'Remaining Stock After Return', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # 寫入數據
    for row_num, (_, row) in enumerate(recommendations_df.iterrows(), 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=row[header])
            cell.border = border
    
    # 調整列寬
    column_widths = [12, 15, 30, 10, 15, 15, 12, 10, 12, 12, 18, 15, 25, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_num).column_letter].width = width
    
    # 工作表 2: 統計摘要
    ws2 = wb.create_sheet("統計摘要")
    
    # KPI 橫幅
    total_recommendations = len(recommendations_df)
    total_return_qty = recommendations_df['Return Qty'].sum() if not recommendations_df.empty else 0
    
    # 分析類型說明
    type_descriptions = {
        "nd_only": "ND 類型退倉分析",
        "rf_only": "RF 類型過剩退倉分析",
        "both": "綜合退貨分析 (ND + RF)"
    }
    analysis_type_desc = type_descriptions.get(calculation_type, "綜合分析")
    
    ws2.cell(row=1, column=1, value="KPI 摘要").font = Font(size=16, bold=True)
    ws2.cell(row=2, column=1, value=f"分析類型: {analysis_type_desc}").font = Font(size=12, bold=True)
    ws2.cell(row=4, column=1, value="總退貨建議數量（條數）:").font = Font(bold=True)
    ws2.cell(row=4, column=2, value=total_recommendations)
    ws2.cell(row=5, column=1, value="總退貨件數:").font = Font(bold=True)
    ws2.cell(row=5, column=2, value=total_return_qty)
    
    # 詳細統計表
    current_row = 8
    
    if not recommendations_df.empty:
        # 按 Article 統計
        ws2.cell(row=current_row, column=1, value="按 Article 統計").font = Font(size=14, bold=True)
        current_row += 2
        
        ws2.cell(row=current_row, column=1, value="Article").font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value="總退貨件數").font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        ws2.cell(row=current_row, column=3, value="涉及OM數量").font = header_font
        ws2.cell(row=current_row, column=3).fill = header_fill
        current_row += 1
        
        article_stats = recommendations_df.groupby('Article').agg({
            'Return Qty': 'sum',
            'OM': 'nunique'
        }).reset_index()
        
        for _, row in article_stats.iterrows():
            ws2.cell(row=current_row, column=1, value=row['Article'])
            ws2.cell(row=current_row, column=2, value=row['Return Qty'])
            ws2.cell(row=current_row, column=3, value=row['OM'])
            current_row += 1
        
        current_row += 2
        
        # 按 OM 統計
        ws2.cell(row=current_row, column=1, value="按 OM 統計").font = Font(size=14, bold=True)
        current_row += 2
        
        ws2.cell(row=current_row, column=1, value="OM").font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value="總退貨件數").font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        ws2.cell(row=current_row, column=3, value="涉及Article數量").font = header_font
        ws2.cell(row=current_row, column=3).fill = header_fill
        current_row += 1
        
        om_stats = recommendations_df.groupby('OM').agg({
            'Return Qty': 'sum',
            'Article': 'nunique'
        }).reset_index()
        
        for _, row in om_stats.iterrows():
            ws2.cell(row=current_row, column=1, value=row['OM'])
            ws2.cell(row=current_row, column=2, value=row['Return Qty'])
            ws2.cell(row=current_row, column=3, value=row['Article'])
            current_row += 1
        
        current_row += 2
        
        # 轉出類型分布
        ws2.cell(row=current_row, column=1, value="轉出類型分布").font = Font(size=14, bold=True)
        current_row += 2
        
        type_stats = recommendations_df.groupby('Type').agg({
            'Return Qty': ['count', 'sum']
        }).round(2)
        type_stats.columns = ['建議數量', '總件數']
        type_stats = type_stats.reset_index()
        
        ws2.cell(row=current_row, column=1, value="類型").font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value="建議數量").font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        ws2.cell(row=current_row, column=3, value="總件數").font = header_font
        ws2.cell(row=current_row, column=3).fill = header_fill
        current_row += 1
        
        for _, row in type_stats.iterrows():
            ws2.cell(row=current_row, column=1, value=row['Type'])
            ws2.cell(row=current_row, column=2, value=row['建議數量'])
            ws2.cell(row=current_row, column=3, value=row['總件數'])
            current_row += 1
        
        current_row += 2
        
        # 退貨前合計統計
        ws2.cell(row=current_row, column=1, value="退貨前合計統計").font = Font(size=14, bold=True)
        current_row += 2
        
        ws2.cell(row=current_row, column=1, value="項目").font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value="合計").font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        current_row += 1
        
        # 計算合計數據
        total_original_stock = df_original['SaSa Net Stock'].sum()
        total_last_month_sold = df_original['Last Month Sold Qty'].sum()
        total_mtd_sold = df_original['MTD Sold Qty'].sum()
        total_safety_stock = df_original['Safety Stock'].sum()
        total_return_qty = recommendations_df['Return Qty'].sum()
        total_remaining_stock = total_original_stock - total_return_qty
        
        ws2.cell(row=current_row, column=1, value="原有存貨")
        ws2.cell(row=current_row, column=2, value=total_original_stock)
        current_row += 1
        
        ws2.cell(row=current_row, column=1, value="上月銷售")
        ws2.cell(row=current_row, column=2, value=total_last_month_sold)
        current_row += 1
        
        ws2.cell(row=current_row, column=1, value="MTD銷售")
        ws2.cell(row=current_row, column=2, value=total_mtd_sold)
        current_row += 1
        
        ws2.cell(row=current_row, column=1, value="Safety QTY")
        ws2.cell(row=current_row, column=2, value=total_safety_stock)
        current_row += 1
        
        ws2.cell(row=current_row, column=1, value="退貨後存貨")
        ws2.cell(row=current_row, column=2, value=total_remaining_stock)
        current_row += 1
        
        current_row += 2
        
        # 退貨類型說明
        ws2.cell(row=current_row, column=1, value="退貨類型說明").font = Font(size=14, bold=True)
        current_row += 2
        
        ws2.cell(row=current_row, column=1, value="類型").font = header_font
        ws2.cell(row=current_row, column=1).fill = header_fill
        ws2.cell(row=current_row, column=2, value="說明").font = header_font
        ws2.cell(row=current_row, column=2).fill = header_fill
        current_row += 1
        
        type_explanations = [
            ['ND', 'ND類型退倉：退回全部現有庫存至D001倉庫。如有銷售記錄，系統會提示 Buyer 需要留意是否需轉成 RF 及設定 Safety Stock'],
            ['RF', 'RF類型過剩退倉：退回過剩庫存（庫存充足且非高銷量店鋪）。若上月銷售量/MTD銷售量 其中一個月 > Safety Qty，退貨後淨餘數量需高於 Safety Qty 的 25% 且至少 +2 件；若上月銷售量/MTD銷售量 同樣地 ≤ Safety Qty，退貨後淨餘數量只需高於 Safety Qty 1 件']
        ]
        
        for explanation in type_explanations:
            ws2.cell(row=current_row, column=1, value=explanation[0])
            ws2.cell(row=current_row, column=2, value=explanation[1])
            current_row += 1
    
    # 調整列寬
    for col in ['A', 'B', 'C']:
        ws2.column_dimensions[col].width = 20
    
    return wb

def quality_check(recommendations_df, original_df):
    """質量檢查"""
    checks = []
    
    if recommendations_df.empty:
        checks.append("✅ 無退貨建議生成")
        return checks
    
    # 檢查 1: Article 和 OM 一致性
    for _, row in recommendations_df.iterrows():
        original_row = original_df[
            (original_df['Article'] == row['Article']) &
            (original_df['Site'] == row['Return Site'])
        ]
        if not original_row.empty and original_row.iloc[0]['OM'] == row['OM']:
            continue
        else:
            checks.append(f"❌ Article {row['Article']} 和 OM {row['OM']} 不一致")
            break
    else:
        checks.append("✅ Article 和 OM 一致性檢查通過")
    
    # 檢查 2: Return Qty 為正整數
    if all(recommendations_df['Return Qty'] > 0):
        checks.append("✅ 所有 Return Qty 為正整數")
    else:
        checks.append("❌ 存在非正整數的 Return Qty")
    
    # 檢查 3: Return Qty 不超過原庫存
    exceeded = False
    for _, row in recommendations_df.iterrows():
        original_row = original_df[
            (original_df['Article'] == row['Article']) &
            (original_df['Site'] == row['Return Site'])
        ]
        if not original_row.empty:
            original_stock = original_row.iloc[0]['SaSa Net Stock']
            if row['Return Qty'] > original_stock:
                exceeded = True
                break
    
    if not exceeded:
        checks.append("✅ Return Qty 不超過原庫存")
    else:
        checks.append("❌ 存在 Return Qty 超過原庫存的情況")
    
    # 檢查 4: Article 格式檢查
    if all(len(str(art)) <= 12 for art in recommendations_df['Article']):
        checks.append("✅ Article 格式正確")
    else:
        checks.append("❌ Article 格式異常")
    
    return checks

def main():
    # 自定義 CSS 樣式
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        text-align: center;
    }
    .info-box {
        background: #f8f9fa;
        border-left: 4px solid #667eea;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .success-box {
        background: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .warning-box {
        background: #fff3cd;
        border-left: 4px solid #ffc107;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .section-header {
        background: #667eea;
        color: white;
        padding: 0.8rem 1.5rem;
        border-radius: 5px;
        margin: 1.5rem 0 1rem 0;
        font-weight: bold;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 主標題
    st.markdown("""
    <div class="main-header">
        <h1 style="margin: 0; font-size: 2.5rem;">📦 退貨建議分析系統</h1>
        <p style="margin: 0.5rem 0 0 0; font-size: 1.1rem; opacity: 0.9;">Return Recommendation Analysis System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # 側邊欄
    st.sidebar.header("🔧 系統設置")
    st.sidebar.markdown("""
    <div class="info-box">
        <strong>接收站點：</strong><br>
        <span style="font-size: 1.2rem; color: #667eea;">D001</span>
    </div>
    """, unsafe_allow_html=True)
    
    # 文件上傳
    st.markdown('<div class="section-header">📤 數據上傳</div>', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "選擇 Excel 文件",
        type=['xlsx'],
        help="支持 .xlsx 格式的 Excel 文件",
        label_visibility="collapsed"
    )
    
    # 處理上傳的文件
    current_file = None
    file_source = ""
    
    if uploaded_file is not None:
        try:
            current_file = pd.read_excel(uploaded_file, dtype={'Article': str})
            file_source = f"上傳文件 ({uploaded_file.name})"
            st.markdown(f"""
            <div class="success-box">
                <strong>✅ 文件上傳成功</strong><br>
                文件名稱: {uploaded_file.name}<br>
                檔案大小: {uploaded_file.size / 1024:.2f} KB
            </div>
            """, unsafe_allow_html=True)
        except Exception as e:
            st.markdown(f"""
            <div class="warning-box">
                <strong>❌ 文件讀取失敗</strong><br>
                錯誤訊息: {str(e)}
            </div>
            """, unsafe_allow_html=True)
    
    if current_file is not None:
        # 數據預覽
        st.markdown('<div class="section-header">🔍 數據預覽</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="info-box">
            <strong>數據來源：</strong> {file_source}
        </div>
        """, unsafe_allow_html=True)
        
        # KPI 卡片
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 2rem; color: #667eea;">{current_file.shape[0]:,}</div>
                <div style="color: #666; margin-top: 0.5rem;">總記錄數</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 2rem; color: #667eea;">{current_file.shape[1]}</div>
                <div style="color: #666; margin-top: 0.5rem;">欄位數</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            nd_count = (current_file['RP Type'] == 'ND').sum()
            rf_count = (current_file['RP Type'] == 'RF').sum()
            st.markdown(f"""
            <div class="metric-card">
                <div style="font-size: 1.5rem; color: #667eea;">ND: {nd_count} | RF: {rf_count}</div>
                <div style="color: #666; margin-top: 0.5rem;">RP Type 分布</div>
            </div>
            """, unsafe_allow_html=True)
        
        # 顯示數據表預覽
        st.markdown('<div class="section-header">📋 數據表預覽 (前 5 行)</div>', unsafe_allow_html=True)
        key_columns = ['Article', 'Article Description', 'OM', 'RP Type', 'Site', 'SaSa Net Stock', 'Pending Received', 'Safety Stock', 'Last Month Sold Qty', 'MTD Sold Qty']
        available_columns = [col for col in key_columns if col in current_file.columns]
        
        if available_columns:
            st.dataframe(current_file[available_columns].head(), use_container_width=True)
        else:
            st.markdown("""
            <div class="warning-box">
                ⚠️ 未找到關鍵欄位，顯示所有欄位前 5 行
            </div>
            """, unsafe_allow_html=True)
            st.dataframe(current_file.head(), use_container_width=True)
        
        # 計算類型選擇
        st.markdown('<div class="section-header">⚙️ 分析設置</div>', unsafe_allow_html=True)
        
        calculation_type = st.radio(
            "選擇計算類型",
            options=[
                ("both", "ND 和 RF 都計算"),
                ("nd_only", "只計算 ND 類型"),
                ("rf_only", "只計算 RF 類型")
            ],
            format_func=lambda x: x[1],
            index=0,
            help="選擇要進行分析的退貨類型",
            label_visibility="visible"
        )
        
        selected_type = calculation_type[0]  # 獲取選中的值
        
        st.markdown("---")
        
        if st.button("🚀 生成退貨建議", type="primary", use_container_width=True):
            with st.spinner("正在處理數據..."):
                # 數據預處理
                processed_df = preprocess_data(current_file)
                
                # 生成退貨建議
                recommendations_df = generate_return_recommendations(processed_df, selected_type)
                
                # 顯示結果
                st.markdown("""
                <div class="success-box">
                    <strong>✅ 分析完成！</strong>
                </div>
                """, unsafe_allow_html=True)
                
                # 基本統計
                st.markdown('<div class="section-header">📊 分析結果</div>', unsafe_allow_html=True)
                
                if not recommendations_df.empty:
                    # 基本統計說明
                    type_description = {
                        "nd_only": "ND 類型退倉",
                        "rf_only": "RF 類型過剩退倉",
                        "both": "綜合退貨分析"
                    }
                    st.markdown(f"""
                    <div class="info-box">
                        <strong>分析類型：</strong> {type_description[selected_type]}
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # KPI 卡片
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div style="font-size: 2rem; color: #667eea;">{len(recommendations_df):,}</div>
                            <div style="color: #666; margin-top: 0.5rem;">退貨建議總數</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col2:
                        st.markdown(f"""
                        <div class="metric-card">
                            <div style="font-size: 2rem; color: #667eea;">{recommendations_df['Return Qty'].sum():,}</div>
                            <div style="color: #666; margin-top: 0.5rem;">總退貨件數</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col3:
                        nd_count = (recommendations_df['Type'] == 'ND').sum()
                        st.markdown(f"""
                        <div class="metric-card">
                            <div style="font-size: 2rem; color: #667eea;">{nd_count}</div>
                            <div style="color: #666; margin-top: 0.5rem;">ND 類型</div>
                        </div>
                        """, unsafe_allow_html=True)
                    with col4:
                        rf_count = (recommendations_df['Type'] == 'RF').sum()
                        st.markdown(f"""
                        <div class="metric-card">
                            <div style="font-size: 2rem; color: #667eea;">{rf_count}</div>
                            <div style="color: #666; margin-top: 0.5rem;">RF 類型</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # 顯示退貨建議表
                    st.markdown('<div class="section-header">🔄 退貨建議表</div>', unsafe_allow_html=True)
                    display_columns = ['Product Hierarchy', 'Article', 'Product Desc', 'OM', 'Return Site', 'Receive Site', 'Return Qty',
                                       'RP Type', 'Stock Qty', 'Safety Qty', 'Last Month Sold Qty', 'MTD Sold Qty',
                                       'Remaining Stock After Return', 'Notes']
                    st.dataframe(recommendations_df[display_columns], use_container_width=True)
                    
                    # 統計圖表
                    st.markdown('<div class="section-header">📈 統計圖表</div>', unsafe_allow_html=True)
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # OM 分布
                        om_stats = recommendations_df.groupby('OM')['Return Qty'].sum().reset_index()
                        st.bar_chart(om_stats.set_index('OM'))
                        st.caption("各 OM 退貨件數分布")
                    
                    with col2:
                        # 類型分布
                        type_stats = recommendations_df.groupby('Type')['Return Qty'].sum().reset_index()
                        st.bar_chart(type_stats.set_index('Type'))
                        st.caption("退貨類型分布")
                    
                else:
                    st.markdown("""
                    <div class="info-box">
                        <strong>📝 未生成任何退貨建議</strong><br><br>
                        <strong>可能原因：</strong><br>
                        • 所有商品均未達到退貨條件<br>
                        • ND 類型商品庫存為 0<br>
                        • RF 類型商品不滿足過剩條件或屬於高銷量商品
                    </div>
                    """, unsafe_allow_html=True)
                
                # 質量檢查
                st.markdown('<div class="section-header">✅ 質量檢查</div>', unsafe_allow_html=True)
                quality_results = quality_check(recommendations_df, processed_df)
                
                for check in quality_results:
                    if "✅" in check:
                        st.markdown(f"""
                        <div class="success-box">
                            {check}
                        </div>
                        """, unsafe_allow_html=True)
                    elif "❌" in check:
                        st.markdown(f"""
                        <div class="warning-box">
                            {check}
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div class="info-box">
                            {check}
                        </div>
                        """, unsafe_allow_html=True)
                
                # 生成並提供下載
                if not recommendations_df.empty:
                    st.markdown('<div class="section-header">💾 下載報告</div>', unsafe_allow_html=True)
                    
                    # 創建 Excel 文件
                    wb = create_excel_report(recommendations_df, processed_df, selected_type)
                    
                    # 生成文件名
                    current_date = datetime.now().strftime("%Y%m%d")
                    filename = f"退貨建議_{current_date}.xlsx"
                    
                    # 保存到內存
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    buffer.seek(0)
                    
                    # 提供下載按鈕
                    st.download_button(
                        label="📥 下載退貨建議報告",
                        data=buffer.getvalue(),
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help=f"下載包含退貨建議和統計摘要的 Excel 文件",
                        use_container_width=True
                    )
                    
                    st.markdown(f"""
                    <div class="success-box">
                        <strong>✅ 報告已準備完成</strong><br>
                        文件名稱: {filename}
                    </div>
                    """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="info-box">
            <strong>👆 請上傳 Excel 文件開始分析</strong>
        </div>
        """, unsafe_allow_html=True)
        
        # 顯示使用說明
        st.markdown('<div class="section-header">📋 使用說明</div>', unsafe_allow_html=True)
        
        with st.expander("💡 系統功能", expanded=True):
            st.markdown("""
            **主要功能：**
            - 📤 支持 Excel 文件上傳
            - 🔍 數據預處理與驗證
            - ⚙️ 自動生成退貨建議（支持 ND 和 RF 類型）
            - 📊 統計分析與圖表展示
            - ✅ 質量檢查與驗證
            - 💾 Excel 報告下載
            """)
        
        with st.expander("🔧 退貨規則說明"):
            st.markdown("""
            **ND 類型退倉：**
            - 適用條件：RP Type = "ND" 且現有庫存 > 0
            - 退貨數量：全部現有庫存退回至 D001 倉庫
            - 特別提示：如有銷售記錄，系統會提示 Buyer 需要留意是否需轉成 RF 及設定 Safety Stock
            - 目的：處理指定需退倉的商品
            
            **RF 類型過剩退倉：**
            - 適用條件：RP Type = "RF"
            - 庫存充足條件：現有庫存 + 在途訂單 > Safety Qty
            - 銷量保護：不屬於該商品的前 20% 高銷量店鋪（避免影響熱銷店鋪）
            - 退貨數量計算：
              - 潛在退貨量 = 總可用庫存 - Safety Qty
              - **若上月銷售量/MTD銷售量 其中一個月 > Safety Qty**：退貨後淨餘數量需高於 Safety Qty 的 25% 且至少 +2 件
              - **若上月銷售量/MTD銷售量 同樣地 ≤ Safety Qty**：退貨後淨餘數量只需高於 Safety Qty 1 件
              - 最終退貨量 = min(潛在退貨量, 總可用庫存 - 最小保留量)
            - 退貨限制：最少退貨 2 件，且不超過現有庫存
            - 目的：優化庫存結構，將過剩庫存退回 D001 倉庫
            """)
        
        with st.expander("📋 必需欄位"):
            st.markdown("""
            **Excel 文件必須包含以下欄位：**
            - Product Hierarchy (產品層級)
            - Article (產品編號)
            - Article Description (產品描述)
            - OM (營運管理單位)
            - RP Type (轉出類型: ND/RF)
            - Site (店鋪編號)
            - SaSa Net Stock (現有庫存)
            - Pending Received (在途訂單)
            - Safety Stock (安全庫存)
            - Last Month Sold Qty (上月銷量)
            - MTD Sold Qty (本月至今銷量)
            """)
    
    # 底部水印
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: #888; font-size: 14px; margin-top: 30px; padding: 20px;'>
            <strong>退貨建議分析系統</strong> | 由 Ricky 開發 | © 2025
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
